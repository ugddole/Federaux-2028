# excel_export.py
# Module réutilisable pour générer des exports Excel stylés (thème FSCF)
# À importer dans app.py : from excel_export import build_export_workbook

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------
# Palette reprise du fichier "Taches_Manifestation_FSCF"
# ---------------------------------------------------------------
COLOR_TITLE_BG = "FFD4006A"     # bandeau titre (rose/magenta FSCF)
COLOR_HEADER_BG = "FF6D006D"    # ligne d'en-têtes de colonnes (violet)
COLOR_WHITE_TEXT = "FFFFFFFF"
COLOR_TEXT = "FF212121"

# Couleurs de fond disponibles pour distinguer des groupes/catégories
# (mêmes teintes pastel que le fichier de référence)
CATEGORY_PALETTE = [
    "FFFCE4EC", "FFF3E5F5", "FFE3F2FD", "FFE8F5E9", "FFFFF3E0",
    "FFE0F7FA", "FFF1F8E9", "FFE8EAF6", "FFFFEBEE", "FFFFF8E1",
    "FFE0F2F1", "FFFBE9E7", "FFF9FBE7",
]

# Couleurs de priorité/statut (mêmes codes que le fichier de référence)
PRIORITY_COLORS = {
    "Impératif": "FFFFCDD2",
    "Important": "FFFFF3E0",
    "Recommandé": "FFE8F5E9",
}


def _category_fill(index):
    """Retourne une couleur de fond stable pour une catégorie donnée (par index)."""
    return CATEGORY_PALETTE[index % len(CATEGORY_PALETTE)]


def build_export_workbook(export_name, headers, rows, category_field=None,
                           priority_field=None, sheet_name=None, export_date=None):
    """
    Construit un classeur Excel stylé selon le modèle FSCF.

    Paramètres
    ----------
    export_name : str
        Nom de l'export tel qu'affiché dans le titre, ex. "Tâches Manifestation",
        "Planning Bénévoles", "Participants".
    headers : list[str]
        Libellés des colonnes, dans l'ordre, ex. ["N°", "Catégorie", "Tâche", ...].
    rows : list[dict]
        Une entrée par ligne de données. Les clés doivent correspondre aux headers
        (en minuscules/slug) — voir l'exemple d'appel en bas de fichier.
    category_field : str, optional
        Clé du dict (dans rows) utilisée pour regrouper/colorer par catégorie,
        comme les sections roses/violettes/bleues du fichier de référence.
    priority_field : str, optional
        Clé du dict utilisée pour appliquer la couleur priorité/statut
        (Impératif = rouge clair, Important = orange clair, Recommandé = vert clair).
    sheet_name : str, optional
        Nom de l'onglet (31 caractères max, imposé par Excel). Par défaut = export_name tronqué.
    export_date : datetime, optional
        Date à afficher dans le titre. Par défaut : maintenant.

    Retourne
    --------
    openpyxl.Workbook
    """
    export_date = export_date or datetime.now()
    date_str = export_date.strftime("%d/%m/%Y %H:%M")
    title = f"{export_name.upper()} — Export du {date_str}"

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or export_name)[:31]

    n_cols = len(headers)
    last_col_letter = get_column_letter(n_cols)

    # --- Ligne 1 : bandeau titre fusionné ---
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=COLOR_WHITE_TEXT)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # --- Ligne 2 : en-têtes de colonnes ---
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color=COLOR_WHITE_TEXT)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 21.75

    # --- Gel des volets sous les en-têtes ---
    ws.freeze_panes = "A3"

    # --- Lignes de données ---
    seen_categories = {}
    keys = list(rows[0].keys()) if rows else []
    for row_idx, row_data in enumerate(rows, start=3):
        # couleur de fond : priorité si dispo, sinon catégorie, sinon blanc
        fill_color = "FFFFFFFF"
        if category_field and category_field in row_data:
            cat = row_data[category_field]
            if cat not in seen_categories:
                seen_categories[cat] = len(seen_categories)
            fill_color = _category_fill(seen_categories[cat])
        if priority_field and priority_field in row_data:
            fill_color = PRIORITY_COLORS.get(row_data[priority_field], fill_color)

        for col_idx, key in enumerate(keys, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(key))
            cell.font = Font(name="Calibri", size=10, color=COLOR_TEXT)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(
                horizontal="left" if col_idx <= 3 else "center",
                vertical="center", wrap_text=False,
            )
        ws.row_dimensions[row_idx].height = 30

    # --- Largeurs de colonnes (auto, avec un plafond raisonnable) ---
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row_data in rows:
            key = keys[col_idx - 1] if col_idx - 1 < len(keys) else None
            if key:
                val = row_data.get(key)
                if val is not None:
                    max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

    ws.sheet_view.showGridLines = False

    # --- Filtre automatique sur les colonnes (ligne d'en-têtes) ---
    last_row = 2 + len(rows) if rows else 2
    ws.auto_filter.ref = f"A2:{last_col_letter}{last_row}"

    return wb


# ---------------------------------------------------------------
# Exemple d'utilisation dans une route Flask (app.py)
# ---------------------------------------------------------------
"""
from flask import send_file
from io import BytesIO
from excel_export import build_export_workbook

@app.route('/export/taches')
def export_taches():
    taches = Tache.query.all()
    rows = [
        {
            "n": t.id,
            "categorie": t.categorie,
            "tache": t.libelle,
            "delai": t.delai,
            "deadline": t.deadline,
            "priorite": t.priorite,
            "responsable": t.responsable,
            "statut": t.statut,
            "notes": t.notes,
        }
        for t in taches
    ]
    headers = ["N°", "Catégorie", "Tâche", "Délai", "Deadline",
               "Priorité", "Responsable", "Statut", "Notes"]

    wb = build_export_workbook(
        export_name="Tâches Manifestation",
        headers=headers,
        rows=rows,
        category_field="categorie",
        priority_field="priorite",
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"taches_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
"""
