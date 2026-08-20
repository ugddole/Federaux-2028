from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, abort
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3, qrcode, os, uuid, csv, json, unicodedata, tempfile, re
from datetime import datetime, timedelta
from io import BytesIO, StringIO
import base64
from fpdf import FPDF
from functools import wraps
from excel_export import build_export_workbook

app = Flask(__name__)
app.secret_key = 'dole2028-fscf-secret-key'

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Veuillez vous connecter.'
login_manager.login_message_category = 'warning'

DATABASE = os.environ.get('DATABASE_PATH', 'dole2028.db')
BADGES_DIR = 'badges'
QR_DIR = os.path.join('static', 'qr')
TMP_DIR  = os.path.join('static', 'tmp')
os.makedirs(BADGES_DIR, exist_ok=True)
os.makedirs(QR_DIR, exist_ok=True)
os.makedirs(TMP_DIR, exist_ok=True)

# ── DATABASE ──────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            nom TEXT NOT NULL,
            prenom TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'participant',
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS participants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            club TEXT NOT NULL,
            region TEXT DEFAULT '',
            categorie TEXT NOT NULL,
            numero_dossard TEXT UNIQUE,
            qr_token TEXT UNIQUE,
            badge_generated INTEGER DEFAULT 0,
            notes TEXT DEFAULT '',
            FOREIGN KEY (user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS benevoles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            telephone TEXT DEFAULT '',
            tshirt TEXT DEFAULT 'M',
            notes TEXT DEFAULT '',
            FOREIGN KEY (user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS missions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            description TEXT DEFAULT '',
            site TEXT NOT NULL,
            jour TEXT NOT NULL,
            heure_debut TEXT NOT NULL,
            heure_fin TEXT NOT NULL,
            nb_places INTEGER DEFAULT 1,
            couleur TEXT DEFAULT '#E2007A'
        );
        CREATE TABLE IF NOT EXISTS affectations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            benevole_id INTEGER NOT NULL,
            mission_id INTEGER NOT NULL,
            UNIQUE(benevole_id, mission_id),
            FOREIGN KEY (benevole_id) REFERENCES benevoles(id),
            FOREIGN KEY (mission_id) REFERENCES missions(id)
        );
        CREATE TABLE IF NOT EXISTS programme (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            categorie TEXT NOT NULL,
            site TEXT NOT NULL,
            jour TEXT NOT NULL,
            heure_debut TEXT NOT NULL,
            heure_fin TEXT NOT NULL,
            nb_gymnastes INTEGER DEFAULT 0,
            notes TEXT DEFAULT '',
            ordre INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS forum_categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            description TEXT DEFAULT '',
            icon TEXT DEFAULT '💬',
            ordre INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS forum_questions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            auteur_id INTEGER NOT NULL,
            categorie_id INTEGER NOT NULL,
            titre TEXT NOT NULL,
            contenu TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            is_resolved INTEGER DEFAULT 0,
            FOREIGN KEY (auteur_id) REFERENCES users(id),
            FOREIGN KEY (categorie_id) REFERENCES forum_categories(id)
        );
        CREATE TABLE IF NOT EXISTS forum_reponses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            question_id INTEGER NOT NULL,
            auteur_id INTEGER NOT NULL,
            contenu TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            is_best INTEGER DEFAULT 0,
            FOREIGN KEY (question_id) REFERENCES forum_questions(id),
            FOREIGN KEY (auteur_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS access_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            participant_id INTEGER,
            scanner_id INTEGER,
            qr_token TEXT,
            timestamp TEXT DEFAULT (datetime('now')),
            site TEXT DEFAULT 'Entrée principale',
            statut TEXT DEFAULT 'ok'
        );
        CREATE TABLE IF NOT EXISTS taches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titre TEXT NOT NULL,
            description TEXT DEFAULT '',
            statut TEXT DEFAULT 'todo',
            priorite TEXT DEFAULT 'normale',
            echeance TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now')),
            created_by INTEGER,
            FOREIGN KEY (created_by) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS sous_taches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tache_id INTEGER NOT NULL,
            titre TEXT NOT NULL,
            statut TEXT DEFAULT 'todo',
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (tache_id) REFERENCES taches(id)
        );
        CREATE TABLE IF NOT EXISTS responsable_categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            categorie TEXT NOT NULL,
            UNIQUE(user_id, categorie),
            FOREIGN KEY (user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS budget_lignes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            categorie TEXT NOT NULL,
            code TEXT DEFAULT '',
            poste TEXT NOT NULL,
            detail TEXT DEFAULT '',
            cout_estime REAL DEFAULT 0,
            subvention_prevue REAL DEFAULT 0,
            reste_a_charge_prevu REAL DEFAULT 0,
            cout_reel REAL DEFAULT 0,
            subvention_recue REAL DEFAULT 0,
            statut TEXT DEFAULT 'prevu',
            notes TEXT DEFAULT '',
            updated_at TEXT DEFAULT (datetime('now'))
        );
    ''')
    conn.commit()
    conn.close()

def migrate_db():
    conn = get_db()
    c = conn.cursor()

    # Migrations
    try:
        c.execute("ALTER TABLE taches ADD COLUMN mission_id INTEGER DEFAULT NULL REFERENCES missions(id)")
        conn.commit()
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE missions ADD COLUMN parent_id INTEGER DEFAULT NULL REFERENCES missions(id)")
        conn.commit()
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE taches ADD COLUMN categorie TEXT DEFAULT ''")
        conn.commit()
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE taches ADD COLUMN responsable_id INTEGER DEFAULT NULL REFERENCES users(id)")
        conn.commit()
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE taches ADD COLUMN delai_libelle TEXT DEFAULT ''")
        conn.commit()
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE taches ADD COLUMN lien TEXT DEFAULT ''")
        conn.commit()
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE users ADD COLUMN droits TEXT DEFAULT ''")
        conn.commit()
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE taches ADD COLUMN lien TEXT DEFAULT ''")
        conn.commit()
    except Exception:
        pass
    try:
         c.execute("ALTER TABLE sous_taches ADD COLUMN lien TEXT DEFAULT ''")
         conn.commit()
    except Exception:
        pass

    if not c.execute("SELECT id FROM users WHERE email='admin@dole2028.fr'").fetchone():
        c.execute("INSERT INTO users (email,password_hash,nom,prenom,role) VALUES (?,?,?,?,?)",
                  ('admin@dole2028.fr', generate_password_hash('admin2028'), 'Admin', 'Dole 2028', 'admin'))
        c.executemany("INSERT INTO forum_categories (nom,description,icon,ordre) VALUES (?,?,?,?)", [
            ('Informations générales', 'Questions générales sur la manifestation', '📋', 1),
            ('Compétition', 'Règlements, épreuves, catégories', '🤸', 2),
            ('Hébergement & Transport', 'Logistique, hôtels, navettes', '🏨', 3),
            ('Bénévoles', 'Organisation, missions, planning', '🙋', 4),
            ('Technique & Matériel', 'Questions techniques sur le site', '⚙️', 5),
        ])
        c.executemany("INSERT INTO missions (nom,description,site,jour,heure_debut,heure_fin,nb_places,couleur) VALUES (?,?,?,?,?,?,?,?)", [
            ('Accueil entrée matin','Accueil des participants à l\'entrée','Dole Expo','Samedi 1er Juillet','07:00','09:00',4,'#E2007A'),
            ('Contrôle accès badges','Vérification QR codes entrée','Dole Expo','Samedi 1er Juillet','07:30','18:00',6,'#0066CC'),
            ('Service restauration midi','Repas de midi participants','Dole Expo','Samedi 1er Juillet','12:00','14:00',8,'#009944'),
            ('Encadrement défilé','Encadrement du défilé centre-ville','Centre-ville','Dimanche 2 Juillet','08:30','11:00',5,'#6a0dad'),
            ('Podium et cérémonie','Gestion du podium remise des prix','Stade Robert Bobin','Dimanche 2 Juillet','10:00','13:30',4,'#009944'),
            ('Accueil associations','Enregistrement des délégations','Dole Expo','Vendredi 30 Juin','15:00','20:00',4,'#E2007A'),
        ])
        c.executemany("INSERT INTO programme (categorie,site,jour,heure_debut,heure_fin,nb_gymnastes,notes,ordre) VALUES (?,?,?,?,?,?,?,?)", [
            ('Poussines / Benjamines','Dole Expo','Samedi 1er Juillet','09:00','10:30',180,'',1),
            ('Minimes Féminines','Dole Expo','Samedi 1er Juillet','09:30','11:30',250,'',2),
            ('Cadettes','Dole Expo','Samedi 1er Juillet','10:30','12:30',220,'',3),
            ('Juniors Féminines','Dole Expo','Samedi 1er Juillet','14:00','15:45',200,'',4),
            ('Séniors Féminines','Dole Expo','Samedi 1er Juillet','15:30','17:00',150,'',5),
            ('Grand Mouvement d\'Ensemble','Stade Robert Bobin','Dimanche 2 Juillet','11:30','12:15',1800,'Toutes catégories',6),
            ('Cérémonie Remise des Prix','Stade Robert Bobin','Dimanche 2 Juillet','12:15','13:15',1800,'',7),
        ])
    conn.commit()
    conn.close()

# ── USER MODEL ────────────────────────────────────────────────────────────────
DROITS_DISPONIBLES = ['participants', 'competition', 'communication', 'organisation', 'administration']

class User(UserMixin):
    def __init__(self, row):
        self.id = row['id']
        self.email = row['email']
        self.nom = row['nom']
        self.prenom = row['prenom']
        self.role = row['role']
        self.droits = (row['droits'] or '').split(',') if row['droits'] else []
    @property
    def is_admin(self): return self.role == 'admin'
    @property
    def is_staff(self): return self.role in ('admin', 'responsable')
    @property
    def full_name(self): return f"{self.prenom} {self.nom}"
    def has_droit(self, droit):
        return self.is_admin or droit in self.droits
    def categories(self):
        if not hasattr(self, '_categories'):
            conn = get_db()
            rows = conn.execute('SELECT categorie FROM responsable_categories WHERE user_id=?', (self.id,)).fetchall()
            conn.close()
            self._categories = [r['categorie'] for r in rows]
        return self._categories

@login_manager.user_loader
def load_user(uid):
    conn = get_db()
    row = conn.execute('SELECT * FROM users WHERE id=?', (uid,)).fetchone()
    conn.close()
    return User(row) if row else None

def can_edit(categorie):
    return current_user.is_admin or categorie in current_user.categories()
app.jinja_env.globals['can_edit'] = can_edit
app.jinja_env.globals['DROITS_DISPONIBLES'] = DROITS_DISPONIBLES

def require_droit(droit):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not current_user.has_droit(droit):
                abort(403)
            return f(*args, **kwargs)
        return wrapped
    return decorator
# ── QR / BADGE ────────────────────────────────────────────────────────────────
def qr_to_base64(token):
    qr = qrcode.QRCode(version=1, box_size=10, border=3)
    qr.add_data(token)
    qr.make(fit=True)
    img = qr.make_image(fill_color='#1a1a2e', back_color='white')
    buf = BytesIO()
    img.save(buf, 'PNG')
    return base64.b64encode(buf.getvalue()).decode()

def generate_badge(p_id):
    conn = get_db()
    p = conn.execute('SELECT p.*,u.nom,u.prenom FROM participants p JOIN users u ON p.user_id=u.id WHERE p.id=?', (p_id,)).fetchone()
    conn.close()
    if not p:
        return None

    qr_path = os.path.join(QR_DIR, f'qr_{p_id}.png')
    qr = qrcode.QRCode(version=1, box_size=8, border=2)
    qr.add_data(p['qr_token'])
    qr.make(fit=True)
    qr.make_image(fill_color='#1a1a2e', back_color='white').save(qr_path)

    pdf = FPDF(orientation='P', unit='mm', format=(86, 125))
    pdf.add_page()
    pdf.set_auto_page_break(False)

    pdf.set_fill_color(226, 0, 122)
    pdf.rect(0, 0, 86, 28, 'F')
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 12)
    pdf.set_xy(3, 5)
    pdf.cell(80, 7, 'DOLE 2028', align='C')
    pdf.set_font('Helvetica', '', 7)
    pdf.set_xy(3, 13)
    pdf.cell(80, 4, 'Manifestation Nationale FSCF', align='C')
    pdf.set_xy(3, 18)
    pdf.cell(80, 4, 'Gymnastique Artistique Feminine', align='C')

    pdf.set_fill_color(26, 26, 46)
    pdf.rect(0, 29, 86, 8, 'F')
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_xy(3, 31)
    pdf.cell(80, 5, str(p['categorie']).upper(), align='C')

    pdf.set_text_color(26, 26, 46)
    pdf.set_font('Helvetica', 'B', 15)
    pdf.set_xy(3, 40)
    pdf.cell(80, 8, str(p['prenom']).upper(), align='C')
    pdf.set_font('Helvetica', 'B', 13)
    pdf.set_xy(3, 49)
    pdf.cell(80, 7, str(p['nom']).upper(), align='C')

    pdf.set_font('Helvetica', '', 9)
    pdf.set_text_color(80, 80, 80)
    pdf.set_xy(3, 58)
    club = str(p['club'])
    if len(club) > 35:
        club = club[:33] + '...'
    pdf.cell(80, 5, club, align='C')

    pdf.set_fill_color(245, 245, 245)
    pdf.rect(3, 65, 80, 8, 'F')
    pdf.set_text_color(226, 0, 122)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_xy(3, 67)
    pdf.cell(80, 5, f"Dossard N° {p['numero_dossard']}", align='C')

    pdf.image(qr_path, x=23, y=75, w=40, h=40)

    pdf.set_text_color(160, 160, 160)
    pdf.set_font('Helvetica', '', 6)
    pdf.set_xy(3, 117)
    pdf.cell(80, 4, 'Scanner ce badge pour valider l\'acces', align='C')

    pdf.set_fill_color(226, 0, 122)
    pdf.rect(0, 121, 86, 4, 'F')

    out = os.path.join(BADGES_DIR, f'badge_{p_id}.pdf')
    pdf.output(out)
    return out

# ── AUTH ──────────────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        conn = get_db()
        row = conn.execute('SELECT * FROM users WHERE email=?', (request.form['email'].strip().lower(),)).fetchone()
        conn.close()
        if row and check_password_hash(row['password_hash'], request.form['password']):
            login_user(User(row), remember=True)
            return redirect(request.args.get('next') or url_for('dashboard'))
        flash('Email ou mot de passe incorrect.', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ── DASHBOARD ─────────────────────────────────────────────────────────────────
@app.route('/')
@login_required
def dashboard():
    conn = get_db()
    stats = {
        'participants': conn.execute('SELECT COUNT(*) FROM participants').fetchone()[0],
        'benevoles':    conn.execute('SELECT COUNT(*) FROM benevoles').fetchone()[0],
        'missions':     conn.execute('SELECT COUNT(*) FROM missions').fetchone()[0],
        'questions':    conn.execute('SELECT COUNT(*) FROM forum_questions').fetchone()[0],
        'badges':       conn.execute('SELECT COUNT(*) FROM participants WHERE badge_generated=1').fetchone()[0],
        'scans':        conn.execute("SELECT COUNT(*) FROM access_logs WHERE date(timestamp)=date('now')").fetchone()[0],
    }
    logs = conn.execute('''
        SELECT al.timestamp, al.site, al.statut, u.nom, u.prenom, p.categorie, p.numero_dossard
        FROM access_logs al
        JOIN participants p ON al.participant_id=p.id
        JOIN users u ON p.user_id=u.id
        ORDER BY al.timestamp DESC LIMIT 8
    ''').fetchall()
    conn.close()
    return render_template('dashboard.html', stats=stats, logs=logs)

# ── PARTICIPANTS ──────────────────────────────────────────────────────────────
@app.route('/participants')
@login_required
@require_droit('participants')
def participants_list():
    conn = get_db()
    q, cat = request.args.get('q',''), request.args.get('cat','')
    sql = 'SELECT p.*,u.nom,u.prenom,u.email FROM participants p JOIN users u ON p.user_id=u.id WHERE 1=1'
    params = []
    if q:
        sql += ' AND (u.nom LIKE ? OR u.prenom LIKE ? OR p.club LIKE ? OR p.numero_dossard LIKE ?)'
        params += [f'%{q}%']*4
    if cat:
        sql += ' AND p.categorie=?'
        params.append(cat)
    sql += ' ORDER BY u.nom,u.prenom'
    items = conn.execute(sql, params).fetchall()
    cats = conn.execute('SELECT DISTINCT categorie FROM participants ORDER BY categorie').fetchall()
    conn.close()
    return render_template('participants_list.html', items=items, cats=cats, q=q, cat=cat)

@app.route('/participants/new', methods=['GET','POST'])
@login_required
def participant_new():
    if not current_user.is_admin:
        abort(403)
    if request.method == 'POST':
        f = request.form
        conn = get_db()
        try:
            conn.execute('INSERT INTO users (email,password_hash,nom,prenom,role) VALUES (?,?,?,?,?)',
                (f['email'].strip().lower(), generate_password_hash(f.get('password','dole2028')),
                 f['nom'].strip(), f['prenom'].strip(), 'participant'))
            uid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            token = str(uuid.uuid4())
            count = conn.execute('SELECT COUNT(*) FROM participants').fetchone()[0]
            dossard = f.get('dossard','').strip() or f'D{2028}{count+1:04d}'
            conn.execute('INSERT INTO participants (user_id,club,region,categorie,numero_dossard,qr_token,notes) VALUES (?,?,?,?,?,?,?)',
                (uid, f['club'].strip(), f.get('region','').strip(), f['categorie'].strip(), dossard, token, f.get('notes','')))
            pid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            conn.commit(); conn.close()
            flash(f'Participant créé — Dossard : {dossard}', 'success')
            return redirect(url_for('participant_detail', id=pid))
        except sqlite3.IntegrityError as e:
            conn.rollback(); conn.close()
            flash('Email ou dossard déjà utilisé.', 'danger')
    return render_template('participant_form.html', action='new', item=None)

@app.route('/participants/<int:id>')
@login_required
def participant_detail(id):
    conn = get_db()
    p = conn.execute('SELECT p.*,u.nom,u.prenom,u.email FROM participants p JOIN users u ON p.user_id=u.id WHERE p.id=?',(id,)).fetchone()
    if not p: abort(404)
    logs = conn.execute('''
        SELECT al.*,su.nom snom,su.prenom sprenom FROM access_logs al
        LEFT JOIN users su ON al.scanner_id=su.id WHERE al.participant_id=? ORDER BY al.timestamp DESC
    ''', (id,)).fetchall()
    conn.close()
    qr = qr_to_base64(p['qr_token'])
    return render_template('participant_detail.html', p=p, qr=qr, logs=logs)

@app.route('/participants/<int:id>/edit', methods=['GET','POST'])
@login_required
def participant_edit(id):
    if not current_user.is_admin: abort(403)
    conn = get_db()
    p = conn.execute('SELECT p.*,u.nom,u.prenom,u.email FROM participants p JOIN users u ON p.user_id=u.id WHERE p.id=?',(id,)).fetchone()
    if not p: abort(404)
    if request.method == 'POST':
        f = request.form
        conn.execute('UPDATE users SET nom=?,prenom=? WHERE id=?',(f['nom'],f['prenom'],p['user_id']))
        conn.execute('UPDATE participants SET club=?,region=?,categorie=?,numero_dossard=?,notes=? WHERE id=?',
            (f['club'],f.get('region',''),f['categorie'],f['dossard'],f.get('notes',''),id))
        conn.commit(); conn.close()
        flash('Participant mis à jour.','success')
        return redirect(url_for('participant_detail', id=id))
    conn.close()
    return render_template('participant_form.html', action='edit', item=p)

@app.route('/participants/<int:id>/delete', methods=['POST'])
@login_required
def participant_delete(id):
    if not current_user.is_admin: return jsonify(error='Non autorisé'), 403
    conn = get_db()
    p = conn.execute('SELECT user_id FROM participants WHERE id=?',(id,)).fetchone()
    if p:
        conn.execute('DELETE FROM participants WHERE id=?',(id,))
        conn.execute('DELETE FROM users WHERE id=?',(p['user_id'],))
        conn.commit()
    conn.close()
    return jsonify(success=True)

@app.route('/participants/<int:id>/badge')
@login_required
def participant_badge(id):
    path = generate_badge(id)
    if not path: abort(404)
    conn = get_db()
    conn.execute('UPDATE participants SET badge_generated=1 WHERE id=?',(id,))
    p = conn.execute('SELECT u.nom,u.prenom FROM participants p JOIN users u ON p.user_id=u.id WHERE p.id=?',(id,)).fetchone()
    conn.commit(); conn.close()
    return send_file(path, as_attachment=True, download_name=f"badge_{p['prenom']}_{p['nom']}.pdf")

# ── BÉNÉVOLES ─────────────────────────────────────────────────────────────────
@app.route('/benevoles')
@login_required
def benevoles_list():
    conn = get_db()
    items = conn.execute('''
        SELECT b.*,u.nom,u.prenom,u.email,u.droits,COUNT(a.id) nb
        FROM benevoles b JOIN users u ON b.user_id=u.id
        LEFT JOIN affectations a ON b.id=a.benevole_id
        GROUP BY b.id ORDER BY u.nom,u.prenom
    ''').fetchall()
    conn.close()
    return render_template('benevoles_list.html', items=items)

@app.route('/benevoles/new', methods=['GET','POST'])
@login_required
def benevole_new():
    if not current_user.is_admin: abort(403)
    if request.method == 'POST':
        f = request.form
        droits = ','.join(request.form.getlist('droits'))
        conn = get_db()
        try:
            conn.execute('INSERT INTO users (email,password_hash,nom,prenom,role,droits) VALUES (?,?,?,?,?,?)',
                (f['email'].strip().lower(), generate_password_hash(f.get('password','benevole2028')),
                 f['nom'].strip(), f['prenom'].strip(), 'benevole', droits))
            uid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            conn.execute('INSERT INTO benevoles (user_id,telephone,tshirt,notes) VALUES (?,?,?,?)',
                (uid, f.get('tel',''), f.get('tshirt','M'), f.get('notes','')))
            conn.commit(); conn.close()
            flash(f"Bénévole {f['prenom']} {f['nom']} créé.", 'success')
            return redirect(url_for('benevoles_list'))
        except sqlite3.IntegrityError:
            conn.rollback(); conn.close()
            flash('Email déjà utilisé.', 'danger')
    return render_template('benevole_form.html', action='new', item=None)

@app.route('/benevoles/<int:id>/edit', methods=['GET','POST'])
@login_required
def benevole_edit(id):
    if not current_user.is_admin: abort(403)
    conn = get_db()
    b = conn.execute('SELECT b.*,u.nom,u.prenom,u.email,u.droits FROM benevoles b JOIN users u ON b.user_id=u.id WHERE b.id=?',(id,)).fetchone()
    if not b: abort(404)
    if request.method == 'POST':
        f = request.form
        droits = ','.join(request.form.getlist('droits'))
        conn.execute('UPDATE users SET nom=?,prenom=?,droits=? WHERE id=?',(f['nom'],f['prenom'],droits,b['user_id']))
        conn.execute('UPDATE benevoles SET telephone=?,tshirt=?,notes=? WHERE id=?',
            (f.get('tel',''),f.get('tshirt','M'),f.get('notes',''),id))
        conn.commit(); conn.close()
        flash('Bénévole mis à jour.','success')
        return redirect(url_for('benevoles_list'))
    conn.close()
    return render_template('benevole_form.html', action='edit', item=b)

def _mission_item(conn, m):
    aff = conn.execute('''
        SELECT b.id,u.nom,u.prenom FROM affectations a
        JOIN benevoles b ON a.benevole_id=b.id
        JOIN users u ON b.user_id=u.id
        WHERE a.mission_id=? ORDER BY u.nom
    ''', (m['id'],)).fetchall()
    taches = conn.execute('''
        SELECT t.*,
            COUNT(st.id) total_st,
            SUM(CASE WHEN st.statut='fait' THEN 1 ELSE 0 END) done_st
        FROM taches t
        LEFT JOIN sous_taches st ON st.tache_id=t.id
        WHERE t.mission_id=?
        GROUP BY t.id
        ORDER BY CASE t.priorite WHEN 'urgente' THEN 0 WHEN 'haute' THEN 1 WHEN 'normale' THEN 2 ELSE 3 END, t.created_at
    ''', (m['id'],)).fetchall()
    return {'m': m, 'aff': aff, 'manque': m['nb_places'] - len(aff), 'taches': taches}

@app.route('/planning')
@login_required
def planning():
    conn = get_db()
    parent_missions = conn.execute(
        'SELECT * FROM missions WHERE parent_id IS NULL ORDER BY jour, heure_debut'
    ).fetchall()
    benevoles = conn.execute(
        'SELECT b.id,u.nom,u.prenom FROM benevoles b JOIN users u ON b.user_id=u.id ORDER BY u.nom'
    ).fetchall()
    data = []
    for m in parent_missions:
        item = _mission_item(conn, m)
        sub_missions = conn.execute(
            'SELECT * FROM missions WHERE parent_id=? ORDER BY heure_debut', (m['id'],)
        ).fetchall()
        item['sous_data'] = [_mission_item(conn, sm) for sm in sub_missions]
        data.append(item)
    conn.close()
    return render_template('planning.html', data=data, benevoles=benevoles)

@app.route('/missions/new', methods=['GET','POST'])
@login_required
def mission_new():
    if not current_user.is_admin: abort(403)
    conn = get_db()
    parent_missions = conn.execute(
        'SELECT * FROM missions WHERE parent_id IS NULL ORDER BY jour, heure_debut'
    ).fetchall()
    if request.method == 'POST':
        f = request.form
        parent_id = f.get('parent_id') or None
        conn.execute(
            'INSERT INTO missions (nom,description,site,jour,heure_debut,heure_fin,nb_places,couleur,parent_id) VALUES (?,?,?,?,?,?,?,?,?)',
            (f['nom'], f.get('desc',''), f['site'], f['jour'], f['hdebut'], f['hfin'],
             int(f.get('places',1)), f.get('couleur','#E2007A'), parent_id))
        conn.commit(); conn.close()
        flash('Mission créée.','success')
        return redirect(url_for('planning'))
    preselect_parent = request.args.get('parent', '')
    conn.close()
    return render_template('mission_form.html', item=None, parent_missions=parent_missions, preselect_parent=preselect_parent)

@app.route('/missions/<int:id>/delete', methods=['POST'])
@login_required
def mission_delete(id):
    if not current_user.is_admin: abort(403)
    conn = get_db()
    sub_ids = [r['id'] for r in conn.execute('SELECT id FROM missions WHERE parent_id=?', (id,)).fetchall()]
    for sid in sub_ids:
        conn.execute('DELETE FROM affectations WHERE mission_id=?', (sid,))
        conn.execute('UPDATE taches SET mission_id=NULL WHERE mission_id=?', (sid,))
    if sub_ids:
        conn.execute('DELETE FROM missions WHERE parent_id=?', (id,))
    conn.execute('DELETE FROM affectations WHERE mission_id=?', (id,))
    conn.execute('UPDATE taches SET mission_id=NULL WHERE mission_id=?', (id,))
    conn.execute('DELETE FROM missions WHERE id=?', (id,))
    conn.commit(); conn.close()
    flash('Mission supprimée.', 'success')
    return redirect(url_for('planning'))

@app.route('/api/affecter', methods=['POST'])
@login_required
def api_affecter():
    if not current_user.is_admin: return jsonify(error='Non autorisé'), 403
    d = request.get_json()
    conn = get_db()
    try:
        if d.get('action') == 'remove':
            conn.execute('DELETE FROM affectations WHERE benevole_id=? AND mission_id=?',(d['bid'],d['mid']))
        else:
            conn.execute('INSERT OR IGNORE INTO affectations (benevole_id,mission_id) VALUES (?,?)',(d['bid'],d['mid']))
        conn.commit(); conn.close()
        return jsonify(success=True)
    except Exception as e:
        conn.close()
        return jsonify(error=str(e)), 500

# ── PROGRAMME ─────────────────────────────────────────────────────────────────
@app.route('/programme')
@login_required
def programme():
    conn = get_db()
    items = conn.execute('SELECT * FROM programme ORDER BY ordre,heure_debut').fetchall()
    conn.close()
    by_day = {}
    for it in items:
        by_day.setdefault(it['jour'], []).append(it)
    return render_template('programme.html', by_day=by_day)

@app.route('/programme/new', methods=['GET','POST'])
@login_required
def programme_new():
    if not current_user.is_admin: abort(403)
    if request.method == 'POST':
        f = request.form
        conn = get_db()
        conn.execute('INSERT INTO programme (categorie,site,jour,heure_debut,heure_fin,nb_gymnastes,notes,ordre) VALUES (?,?,?,?,?,?,?,?)',
            (f['cat'],f['site'],f['jour'],f['hdebut'],f['hfin'],int(f.get('nb',0)),f.get('notes',''),int(f.get('ordre',0))))
        conn.commit(); conn.close()
        flash('Créneau ajouté.','success')
        return redirect(url_for('programme'))
    return render_template('programme_form.html', item=None)

@app.route('/programme/<int:id>/edit', methods=['GET','POST'])
@login_required
def programme_edit(id):
    if not current_user.is_admin: abort(403)
    conn = get_db()
    item = conn.execute('SELECT * FROM programme WHERE id=?',(id,)).fetchone()
    if not item: abort(404)
    if request.method == 'POST':
        f = request.form
        conn.execute('UPDATE programme SET categorie=?,site=?,jour=?,heure_debut=?,heure_fin=?,nb_gymnastes=?,notes=?,ordre=? WHERE id=?',
            (f['cat'],f['site'],f['jour'],f['hdebut'],f['hfin'],int(f.get('nb',0)),f.get('notes',''),int(f.get('ordre',0)),id))
        conn.commit(); conn.close()
        flash('Créneau mis à jour.','success')
        return redirect(url_for('programme'))
    conn.close()
    return render_template('programme_form.html', item=item)

@app.route('/programme/<int:id>/delete', methods=['POST'])
@login_required
def programme_del(id):
    if not current_user.is_admin: abort(403)
    conn = get_db()
    conn.execute('DELETE FROM programme WHERE id=?',(id,))
    conn.commit(); conn.close()
    flash('Créneau supprimé.','success')
    return redirect(url_for('programme'))

# ── FORUM ─────────────────────────────────────────────────────────────────────
@app.route('/forum')
@login_required
def forum():
    conn = get_db()
    cats = conn.execute('''
        SELECT fc.*,COUNT(fq.id) nb,SUM(CASE WHEN fq.is_resolved=0 THEN 1 ELSE 0 END) ouvertes
        FROM forum_categories fc LEFT JOIN forum_questions fq ON fc.id=fq.categorie_id
        GROUP BY fc.id ORDER BY fc.ordre
    ''').fetchall()
    recent = conn.execute('''
        SELECT fq.*,u.nom,u.prenom,fc.nom cat_nom,COUNT(fr.id) nb_rep
        FROM forum_questions fq JOIN users u ON fq.auteur_id=u.id
        JOIN forum_categories fc ON fq.categorie_id=fc.id
        LEFT JOIN forum_reponses fr ON fq.id=fr.question_id
        GROUP BY fq.id ORDER BY fq.created_at DESC LIMIT 6
    ''').fetchall()
    conn.close()
    return render_template('forum_index.html', cats=cats, recent=recent)

@app.route('/forum/cat/<int:id>')
@login_required
def forum_cat(id):
    conn = get_db()
    cat = conn.execute('SELECT * FROM forum_categories WHERE id=?',(id,)).fetchone()
    if not cat: abort(404)
    qs = conn.execute('''
        SELECT fq.*,u.nom,u.prenom,COUNT(fr.id) nb FROM forum_questions fq
        JOIN users u ON fq.auteur_id=u.id
        LEFT JOIN forum_reponses fr ON fq.id=fr.question_id
        WHERE fq.categorie_id=? GROUP BY fq.id ORDER BY fq.created_at DESC
    ''', (id,)).fetchall()
    conn.close()
    return render_template('forum_cat.html', cat=cat, qs=qs)

@app.route('/forum/q/<int:id>', methods=['GET','POST'])
@login_required
def forum_question(id):
    conn = get_db()
    q = conn.execute('''
        SELECT fq.*,u.nom,u.prenom,fc.nom cat_nom FROM forum_questions fq
        JOIN users u ON fq.auteur_id=u.id
        JOIN forum_categories fc ON fq.categorie_id=fc.id WHERE fq.id=?
    ''', (id,)).fetchone()
    if not q: abort(404)
    if request.method == 'POST':
        txt = request.form.get('contenu','').strip()
        if txt:
            conn.execute('INSERT INTO forum_reponses (question_id,auteur_id,contenu) VALUES (?,?,?)',(id,current_user.id,txt))
            conn.commit()
            flash('Réponse publiée.','success')
    reps = conn.execute('''
        SELECT fr.*,u.nom,u.prenom,u.role FROM forum_reponses fr
        JOIN users u ON fr.auteur_id=u.id WHERE fr.question_id=? ORDER BY fr.created_at
    ''', (id,)).fetchall()
    conn.close()
    return render_template('forum_question.html', q=q, reps=reps)

@app.route('/forum/ask', methods=['GET','POST'])
@login_required
def forum_ask():
    conn = get_db()
    cats = conn.execute('SELECT * FROM forum_categories ORDER BY ordre').fetchall()
    if request.method == 'POST':
        f = request.form
        if f.get('titre') and f.get('contenu') and f.get('cat'):
            conn.execute('INSERT INTO forum_questions (auteur_id,categorie_id,titre,contenu) VALUES (?,?,?,?)',
                (current_user.id, f['cat'], f['titre'].strip(), f['contenu'].strip()))
            qid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            conn.commit(); conn.close()
            flash('Question publiée.','success')
            return redirect(url_for('forum_question', id=qid))
        flash('Veuillez remplir tous les champs.','danger')
    conn.close()
    return render_template('forum_ask.html', cats=cats)

@app.route('/forum/q/<int:id>/resolve', methods=['POST'])
@login_required
def forum_resolve(id):
    conn = get_db()
    q = conn.execute('SELECT * FROM forum_questions WHERE id=?',(id,)).fetchone()
    if q and (current_user.is_admin or q['auteur_id'] == current_user.id):
        conn.execute('UPDATE forum_questions SET is_resolved=? WHERE id=?',(0 if q['is_resolved'] else 1, id))
        conn.commit()
    conn.close()
    return redirect(url_for('forum_question', id=id))

# ── SCANNER ───────────────────────────────────────────────────────────────────
@app.route('/scanner')
@login_required
def scanner():
    conn = get_db()
    logs = conn.execute('''
        SELECT al.*,u.nom,u.prenom,p.categorie,p.numero_dossard,p.club
        FROM access_logs al JOIN participants p ON al.participant_id=p.id
        JOIN users u ON p.user_id=u.id ORDER BY al.timestamp DESC LIMIT 20
    ''').fetchall()
    conn.close()
    return render_template('scanner.html', logs=logs)

@app.route('/api/scan', methods=['POST'])
@login_required
def api_scan():
    d = request.get_json()
    token = (d.get('token') or '').strip()
    site  = d.get('site', 'Entrée principale')
    conn = get_db()
    p = conn.execute('''
        SELECT p.*,u.nom,u.prenom FROM participants p
        JOIN users u ON p.user_id=u.id WHERE p.qr_token=?
    ''', (token,)).fetchone()
    if not p:
        conn.close()
        return jsonify(statut='invalide', message='QR code non reconnu'), 404
    already = conn.execute("SELECT id FROM access_logs WHERE participant_id=? AND date(timestamp)=date('now')",(p['id'],)).fetchone()
    statut = 'deja_scanne' if already else 'ok'
    conn.execute('INSERT INTO access_logs (participant_id,scanner_id,qr_token,site,statut) VALUES (?,?,?,?,?)',
        (p['id'], current_user.id, token, site, statut))
    conn.commit(); conn.close()
    return jsonify(statut=statut, nom=p['nom'], prenom=p['prenom'],
                   club=p['club'], categorie=p['categorie'], dossard=p['numero_dossard'])

# ── ADMIN ─────────────────────────────────────────────────────────────────────
@app.route('/admin')
@login_required
def admin():
    if not current_user.is_admin: abort(403)
    conn = get_db()
    users = conn.execute('SELECT * FROM users ORDER BY role,nom').fetchall()
    responsables = conn.execute("SELECT * FROM users WHERE role='responsable' ORDER BY nom").fetchall()
    resp_cats = {}
    for u in responsables:
        rows = conn.execute('SELECT categorie FROM responsable_categories WHERE user_id=?', (u['id'],)).fetchall()
        resp_cats[u['id']] = [r['categorie'] for r in rows]
    cats_taches = [r['categorie'] for r in conn.execute(
        "SELECT DISTINCT categorie FROM taches WHERE categorie!='' ORDER BY categorie").fetchall()]
    cats_budget = [r['categorie'] for r in conn.execute(
        "SELECT DISTINCT categorie FROM budget_lignes ORDER BY categorie").fetchall()]
    all_cats = sorted(set(cats_taches) | set(cats_budget))
    conn.close()
    return render_template('admin.html', users=users, responsables=responsables, resp_cats=resp_cats, all_cats=all_cats)

@app.route('/admin/responsable/new', methods=['POST'])
@login_required
def admin_responsable_new():
    if not current_user.is_admin: abort(403)
    f = request.form
    conn = get_db()
    try:
        conn.execute('INSERT INTO users (email,password_hash,nom,prenom,role) VALUES (?,?,?,?,?)',
            (f['email'].strip().lower(), generate_password_hash(f.get('password','dole2028')),
             f['nom'].strip(), f['prenom'].strip(), 'responsable'))
        conn.commit()
        flash(f"Responsable {f['prenom']} {f['nom']} créé.", 'success')
    except sqlite3.IntegrityError:
        conn.rollback()
        flash('Email déjà utilisé.', 'danger')
    conn.close()
    return redirect(url_for('admin'))

@app.route('/admin/responsable/<int:id>/categories', methods=['POST'])
@login_required
def admin_responsable_categories(id):
    if not current_user.is_admin: abort(403)
    conn = get_db()
    u = conn.execute("SELECT * FROM users WHERE id=? AND role='responsable'", (id,)).fetchone()
    if not u:
        conn.close(); abort(404)
    cats = [c.strip() for c in request.form.getlist('categories') if c.strip()]
    conn.execute('DELETE FROM responsable_categories WHERE user_id=?', (id,))
    for c in cats:
        conn.execute('INSERT OR IGNORE INTO responsable_categories (user_id,categorie) VALUES (?,?)', (id, c))
    conn.commit(); conn.close()
    flash(f"Catégories mises à jour pour {u['prenom']} {u['nom']}.", 'success')
    return redirect(url_for('admin'))

@app.route('/admin/user/<int:id>/reset', methods=['POST'])
@login_required
def admin_reset(id):
    if not current_user.is_admin: abort(403)
    pwd = request.form.get('password','dole2028')
    conn = get_db()
    conn.execute('UPDATE users SET password_hash=? WHERE id=?',(generate_password_hash(pwd),id))
    conn.commit(); conn.close()
    flash('Mot de passe réinitialisé.','success')
    return redirect(url_for('admin'))

@app.route('/admin/user/<int:id>/delete', methods=['POST'])
@login_required
def admin_user_delete(id):
    if not current_user.is_admin: abort(403)
    if id == current_user.id:
        flash('Vous ne pouvez pas supprimer votre propre compte.','danger')
        return redirect(url_for('admin'))
    conn = get_db()
    conn.execute('DELETE FROM users WHERE id=?',(id,))
    conn.commit(); conn.close()
    flash('Utilisateur supprimé.','success')
    return redirect(url_for('admin'))

# ── IMPORT ADAGIO ─────────────────────────────────────────────────────────────

def _normalize(s):
    """Remove accents and lowercase for email generation."""
    return ''.join(c for c in unicodedata.normalize('NFD', str(s))
                   if unicodedata.category(c) != 'Mn').lower()

def _find_col(headers_upper, candidates):
    """Find a column by trying several candidate names (case-insensitive partial match)."""
    for cand in candidates:
        for i, h in enumerate(headers_upper):
            if cand in h or h in cand:
                return i
    return None

def parse_adagio_file(file_storage):
    """
    Parse a CSV or Excel export from ADAGIO.
    Returns (headers, raw_rows_as_dicts).
    """
    fname = file_storage.filename.lower()

    if fname.endswith('.csv'):
        raw = file_storage.read().decode('utf-8-sig', errors='replace')
        # Detect delimiter: semicolon is standard in French exports
        delim = ';' if raw.count(';') >= raw.count(',') else ','
        reader = csv.DictReader(StringIO(raw), delimiter=delim)
        headers = list(reader.fieldnames or [])
        rows = [dict(r) for r in reader if any(v and str(v).strip() for v in r.values())]

    elif fname.endswith(('.xlsx', '.xls')):
        import openpyxl
        wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
        ws = wb.active
        data = [row for row in ws.iter_rows(values_only=True)]
        if not data:
            return [], []
        headers = [str(c).strip() if c is not None else f'Col{i}' for i, c in enumerate(data[0])]
        rows = []
        for row in data[1:]:
            if any(v is not None for v in row):
                rows.append({headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)})
    else:
        raise ValueError('Format non supporté. Utilisez CSV (.csv) ou Excel (.xlsx).')

    return headers, rows


def map_adagio_rows(headers, rows):
    """
    Auto-map ADAGIO columns to internal fields.
    Returns list of dicts with keys: nom, prenom, club, region, categorie, email, licence.
    """
    H = [h.upper().strip() for h in headers]

    col_idx = {
        'nom':       _find_col(H, ['NOM', 'LAST_NAME', 'GYMNASTE_NOM']),
        'prenom':    _find_col(H, ['PRENOM', 'PRÉNOM', 'FIRST_NAME', 'GYMNASTE_PRENOM']),
        'club':      _find_col(H, ['CLUB', 'ASSOCIATION', 'STRUCT', 'LIBELLE_ASSO', 'LIBASSOC']),
        'region':    _find_col(H, ['REGION', 'LIGUE', 'COMITE', 'COMITÉ', 'LIBLIGUE', 'LIBCOMITE']),
        'categorie': _find_col(H, ['CATEG', 'CATEGORY', 'DIVISION', 'NIVEAU']),
        'email':     _find_col(H, ['EMAIL', 'MAIL', 'COURRIEL', 'E-MAIL']),
        'licence':   _find_col(H, ['LICENCE', 'LICENSE', 'NUMLICENCE', 'N_LICENCE']),
    }

    # Build readable mapping for display (col name → field)
    col_names = {field: (headers[idx] if idx is not None else None) for field, idx in col_idx.items()}

    mapped = []
    for row in rows:
        def get(field):
            idx = col_idx.get(field)
            if idx is None:
                return ''
            return list(row.values())[idx] if idx < len(row) else row.get(headers[idx], '')

        nom      = get('nom').strip().upper()
        prenom   = get('prenom').strip().capitalize()
        club     = get('club').strip() or 'Non renseigné'
        region   = get('region').strip()
        categorie = get('categorie').strip() or 'Non renseignée'
        email    = get('email').strip().lower()
        licence  = get('licence').strip()

        if not nom or not prenom:
            continue

        # Auto-generate email from name if missing
        if not email:
            p = _normalize(prenom).replace(' ', '').replace('-', '')
            n = _normalize(nom).replace(' ', '').replace('-', '')
            email = f"{p}.{n}@dole2028.fr"

        mapped.append(dict(nom=nom, prenom=prenom, club=club, region=region,
                           categorie=categorie, email=email, licence=licence))

    return mapped, col_names


@app.route('/import', methods=['GET', 'POST'])
@login_required
def import_adagio():
    if not current_user.is_admin:
        abort(403)

    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename:
            flash('Aucun fichier sélectionné.', 'danger')
            return redirect(request.url)
        try:
            headers, raw_rows = parse_adagio_file(f)
            mapped, col_names = map_adagio_rows(headers, raw_rows)

            if not mapped:
                flash('Aucune ligne exploitable trouvée dans le fichier. Vérifiez le format.', 'warning')
                return redirect(request.url)

            # Check duplicates against existing DB
            conn = get_db()
            existing_emails = {r[0] for r in conn.execute('SELECT email FROM users').fetchall()}
            conn.close()
            for r in mapped:
                r['doublon'] = (r['email'] in existing_emails)

            # Store full dataset in a temp JSON file
            tmp_path = os.path.join(TMP_DIR, f'adagio_{uuid.uuid4().hex}.json')
            with open(tmp_path, 'w', encoding='utf-8') as fp:
                json.dump(mapped, fp, ensure_ascii=False)

            nb_doublons = sum(1 for r in mapped if r['doublon'])
            return render_template('import_preview.html',
                                   rows=mapped[:50], total=len(mapped),
                                   nb_doublons=nb_doublons,
                                   col_names=col_names,
                                   tmp_file=os.path.basename(tmp_path))
        except Exception as e:
            flash(f'Erreur de lecture : {e}', 'danger')

    return render_template('import_adagio.html')


@app.route('/import/confirm', methods=['POST'])
@login_required
def import_confirm():
    if not current_user.is_admin:
        abort(403)

    tmp_name = request.form.get('tmp_file', '')
    tmp_path = os.path.join(TMP_DIR, tmp_name)

    if not tmp_name or not os.path.exists(tmp_path):
        flash('Session expirée. Veuillez relancer l\'import.', 'danger')
        return redirect(url_for('import_adagio'))

    with open(tmp_path, encoding='utf-8') as fp:
        rows = json.load(fp)
    os.unlink(tmp_path)

    skip_doublons = request.form.get('skip_doublons') == '1'
    conn = get_db()
    ok, skip, err = 0, 0, 0

    for r in rows:
        if r.get('doublon') and skip_doublons:
            skip += 1
            continue
        try:
            conn.execute('INSERT INTO users (email,password_hash,nom,prenom,role) VALUES (?,?,?,?,?)',
                (r['email'], generate_password_hash('dole2028'), r['nom'], r['prenom'], 'participant'))
            uid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            token = str(uuid.uuid4())
            count = conn.execute('SELECT COUNT(*) FROM participants').fetchone()[0]
            dossard = f"D{2028}{count + 1:04d}"
            conn.execute('INSERT INTO participants (user_id,club,region,categorie,numero_dossard,qr_token) VALUES (?,?,?,?,?,?)',
                (uid, r['club'], r['region'], r['categorie'], dossard, token))
            conn.commit()
            ok += 1
        except sqlite3.IntegrityError:
            conn.rollback()
            skip += 1
        except Exception:
            conn.rollback()
            err += 1

    conn.close()
    cat = 'success' if ok > 0 else 'warning'
    flash(f'✅ Import terminé — {ok} gymnaste(s) créé(s), {skip} doublon(s) ignoré(s), {err} erreur(s).', cat)
    return redirect(url_for('participants_list'))


# ── TÂCHES ───────────────────────────────────────────────────────────────────
def _tache_categories(conn):
    return [r['categorie'] for r in conn.execute(
        "SELECT DISTINCT categorie FROM taches WHERE categorie!='' ORDER BY categorie").fetchall()]

def _editable_categories(conn):
    """Categories the current user may create/edit tasks in."""
    if current_user.is_admin:
        return _tache_categories(conn)
    return current_user.categories()
    
@app.route('/taches/urgentes')
@login_required
def taches_urgentes():
    if not current_user.is_staff: abort(403)
    conn = get_db()
    taches = conn.execute('''
        SELECT t.*, m.nom mission_nom, m.couleur mission_couleur, u.nom resp_nom, u.prenom resp_prenom
        FROM taches t
        LEFT JOIN missions m ON t.mission_id = m.id
        LEFT JOIN users u ON t.responsable_id = u.id
        WHERE t.echeance != '' AND t.statut != 'termine'
        ORDER BY t.echeance
    ''').fetchall()
    conn.close()

    today = datetime.now().date()
    limite = today + timedelta(days=60)

    en_retard, a_venir = [], []
    for t in taches:
        try:
            d = datetime.strptime(t['echeance'], '%Y-%m-%d').date()
        except (ValueError, TypeError):
            continue
        if d < today:
            en_retard.append(t)
        elif d <= limite:
            a_venir.append(t)

    return render_template('taches_urgentes.html', en_retard=en_retard, a_venir=a_venir, today=today)

@app.route('/taches')
@login_required
def taches_list():
    if not current_user.is_staff: abort(403)
    conn = get_db()
    taches = conn.execute('''
        SELECT t.*, m.nom mission_nom, m.couleur mission_couleur, u.nom resp_nom, u.prenom resp_prenom
        FROM taches t
        LEFT JOIN missions m ON t.mission_id = m.id
        LEFT JOIN users u ON t.responsable_id = u.id
        ORDER BY t.categorie, t.created_at
    ''').fetchall()
    by_cat = {}
    for t in taches:
        total = conn.execute('SELECT COUNT(*) FROM sous_taches WHERE tache_id=?', (t['id'],)).fetchone()[0]
        done  = conn.execute("SELECT COUNT(*) FROM sous_taches WHERE tache_id=? AND statut='fait'", (t['id'],)).fetchone()[0]
        cat = t['categorie'] or 'Sans catégorie'
        by_cat.setdefault(cat, []).append({'t': t, 'total': total, 'done': done, 'can_edit': can_edit(t['categorie'])})
    conn.close()
    return render_template('taches_list.html', by_cat=by_cat)

@app.route('/taches/new', methods=['GET','POST'])
@login_required
def tache_new():
    if not current_user.is_staff: abort(403)
    conn = get_db()
    missions = conn.execute('SELECT * FROM missions ORDER BY jour, heure_debut').fetchall()
    responsables = conn.execute("SELECT * FROM users WHERE role IN ('admin','responsable') ORDER BY nom").fetchall()
    editable_cats = _editable_categories(conn)
    if request.method == 'POST':
        f = request.form
        categorie = f.get('categorie','').strip()
        if not can_edit(categorie):
            conn.close(); abort(403)
        mid = f.get('mission_id') or None
        rid = f.get('responsable_id') or None
        conn.execute('''INSERT INTO taches
            (titre,description,statut,priorite,echeance,mission_id,created_by,categorie,responsable_id,delai_libelle,lien)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
            (f['titre'].strip(), f.get('description','').strip(),
             f.get('statut','todo'), f.get('priorite','normale'),
             f.get('echeance',''), mid, current_user.id,
             categorie, rid, f.get('delai_libelle','').strip(),
             f.get('lien','').strip()))
        tid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.commit(); conn.close()
        flash('Tâche créée.', 'success')
        return redirect(url_for('tache_detail', id=tid))
    preselect_mission = request.args.get('mission', '')
    conn.close()
    return render_template('tache_form.html', action='new', item=None, missions=missions,
                           responsables=responsables, editable_cats=editable_cats, preselect_mission=preselect_mission)

@app.route('/taches/<int:id>')
@login_required
def tache_detail(id):
    if not current_user.is_staff: abort(403)
    conn = get_db()
    t = conn.execute('''
        SELECT t.*, u.nom resp_nom, u.prenom resp_prenom
        FROM taches t LEFT JOIN users u ON t.responsable_id = u.id WHERE t.id=?''', (id,)).fetchone()
    if not t: abort(404)
    sous_taches = conn.execute('SELECT * FROM sous_taches WHERE tache_id=? ORDER BY created_at', (id,)).fetchall()
    mission = conn.execute('SELECT * FROM missions WHERE id=?', (t['mission_id'],)).fetchone() if t['mission_id'] else None
    conn.close()
    total = len(sous_taches)
    done  = sum(1 for s in sous_taches if s['statut'] == 'fait')
    editable = can_edit(t['categorie'])
    return render_template('tache_detail.html', t=t, sous_taches=sous_taches, total=total, done=done,
                           mission=mission, editable=editable)

@app.route('/taches/<int:id>/edit', methods=['GET','POST'])
@login_required
def tache_edit(id):
    if not current_user.is_staff: abort(403)
    conn = get_db()
    t = conn.execute('SELECT * FROM taches WHERE id=?', (id,)).fetchone()
    if not t: abort(404)
    if not can_edit(t['categorie']):
        conn.close(); abort(403)
    missions = conn.execute('SELECT * FROM missions ORDER BY jour, heure_debut').fetchall()
    responsables = conn.execute("SELECT * FROM users WHERE role IN ('admin','responsable') ORDER BY nom").fetchall()
    editable_cats = _editable_categories(conn)
    if request.method == 'POST':
        f = request.form
        categorie = f.get('categorie','').strip()
        if not can_edit(categorie):
            conn.close(); abort(403)
        mid = f.get('mission_id') or None
        rid = f.get('responsable_id') or None
        conn.execute('''UPDATE taches SET titre=?,description=?,statut=?,priorite=?,echeance=?,mission_id=?,
            categorie=?,responsable_id=?,delai_libelle=?,lien=? WHERE id=?''',
            (f['titre'].strip(), f.get('description','').strip(),
             f.get('statut','todo'), f.get('priorite','normale'),
             f.get('echeance',''), mid, categorie, rid, f.get('delai_libelle','').strip(),
             f.get('lien','').strip(), id))
        conn.commit(); conn.close()
        flash('Tâche mise à jour.', 'success')
        return redirect(url_for('tache_detail', id=id))
    conn.close()
    return render_template('tache_form.html', action='edit', item=t, missions=missions,
                           responsables=responsables, editable_cats=editable_cats, preselect_mission='')

@app.route('/taches/<int:id>/delete', methods=['POST'])
@login_required
def tache_delete(id):
    if not current_user.is_staff: abort(403)
    conn = get_db()
    t = conn.execute('SELECT categorie FROM taches WHERE id=?', (id,)).fetchone()
    if not t: conn.close(); abort(404)
    if not can_edit(t['categorie']):
        conn.close(); abort(403)
    conn.execute('DELETE FROM sous_taches WHERE tache_id=?', (id,))
    conn.execute('DELETE FROM taches WHERE id=?', (id,))
    conn.commit(); conn.close()
    flash('Tâche supprimée.', 'success')
    return redirect(url_for('taches_list'))

@app.route('/taches/<int:id>/sous-tache', methods=['POST'])
@login_required
def sous_tache_new(id):
    if not current_user.is_staff: abort(403)
    conn = get_db()
    t = conn.execute('SELECT categorie FROM taches WHERE id=?', (id,)).fetchone()
    if not t: conn.close(); abort(404)
    if not can_edit(t['categorie']):
        conn.close(); abort(403)
    titre = request.form.get('titre','').strip()
    lien = request.form.get('lien','').strip()
    if titre:
        conn.execute('INSERT INTO sous_taches (tache_id,titre,lien) VALUES (?,?,?)', (id, titre, lien))
        conn.commit()
    conn.close()
    return redirect(url_for('tache_detail', id=id))
    
@app.route('/sous-tache/<int:id>/toggle', methods=['POST'])
@login_required
def sous_tache_toggle(id):
    if not current_user.is_staff: abort(403)
    conn = get_db()
    st = conn.execute('''SELECT st.*, t.categorie FROM sous_taches st
        JOIN taches t ON st.tache_id = t.id WHERE st.id=?''', (id,)).fetchone()
    if not st: conn.close(); abort(404)
    if not can_edit(st['categorie']):
        conn.close(); abort(403)
    new_statut = 'todo' if st['statut'] == 'fait' else 'fait'
    conn.execute('UPDATE sous_taches SET statut=? WHERE id=?', (new_statut, id))
    conn.commit()
    tache_id = st['tache_id']
    conn.close()
    return redirect(url_for('tache_detail', id=tache_id))
    
@app.route('/sous-tache/<int:id>/edit', methods=['POST'])
@login_required
def sous_tache_edit(id):
    if not current_user.is_staff: abort(403)
    conn = get_db()
    st = conn.execute('''SELECT st.*, t.categorie FROM sous_taches st
        JOIN taches t ON st.tache_id = t.id WHERE st.id=?''', (id,)).fetchone()
    if not st: conn.close(); abort(404)
    if not can_edit(st['categorie']):
        conn.close(); abort(403)
    titre = request.form.get('titre','').strip()
    lien = request.form.get('lien','').strip()
    tache_id = st['tache_id']
    if titre:
        conn.execute('UPDATE sous_taches SET titre=?, lien=? WHERE id=?', (titre, lien, id))
        conn.commit()
    conn.close()
    return redirect(url_for('tache_detail', id=tache_id))

@app.route('/sous-tache/<int:id>/delete', methods=['POST'])
@login_required
def sous_tache_delete(id):
    if not current_user.is_staff: abort(403)
    conn = get_db()
    st = conn.execute('''SELECT st.*, t.categorie FROM sous_taches st
        JOIN taches t ON st.tache_id = t.id WHERE st.id=?''', (id,)).fetchone()
    if not st: conn.close(); abort(404)
    if not can_edit(st['categorie']):
        conn.close(); abort(403)
    tache_id = st['tache_id']
    conn.execute('DELETE FROM sous_taches WHERE id=?', (id,))
    conn.commit(); conn.close()
    return redirect(url_for('tache_detail', id=tache_id))

# ── IMPORT TÂCHES ─────────────────────────────────────────────────────────────
STATUT_MAP = {
    'a faire': 'todo', 'à faire': 'todo',
    'en cours': 'en_cours',
    'termine': 'termine', 'terminé': 'termine', 'fait': 'termine',
}

def parse_taches_file(file_storage):
    import openpyxl
    wb = openpyxl.load_workbook(file_storage, data_only=True)
    ws = None
    for s in wb.worksheets:
        if 'tach' in _normalize(s.title):
            ws = s
            break
    if ws is None:
        ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    data = []
    for row in rows[2:]:
        if not row or row[0] is None or not isinstance(row[0], (int, float)):
            continue  # blank line or section header ("▶ CATEGORIE")
        categorie = str(row[1]).strip() if row[1] else ''
        titre = str(row[2]).strip() if row[2] else ''
        if not titre:
            continue
        delai_libelle = str(row[3]).strip() if row[3] else ''
        deadline = row[4] if len(row) > 4 else None
        if isinstance(deadline, datetime):
            echeance = deadline.strftime('%Y-%m-%d')
        elif deadline:
            echeance = str(deadline).strip()
        else:
            echeance = ''
        priorite = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        responsable_txt = str(row[6]).strip() if len(row) > 6 and row[6] else ''
        statut_txt = _normalize(row[7]).strip() if len(row) > 7 and row[7] else ''
        statut = STATUT_MAP.get(statut_txt, 'todo')
        lien = str(row[8]).strip() if len(row) > 8 and row[8] else ''
        notes = str(row[9]).strip() if len(row) > 9 and row[9] else ''
        data.append(dict(categorie=categorie, titre=titre, delai_libelle=delai_libelle,
                          echeance=echeance, priorite=priorite, responsable_txt=responsable_txt,
                          statut=statut, lien=lien, notes=notes))
    return data

def _match_responsable(conn, txt):
    if not txt:
        return None
    t = _normalize(txt)
    for u in conn.execute("SELECT id,nom,prenom FROM users WHERE role IN ('admin','responsable')").fetchall():
        if _normalize(u['nom']) in t or _normalize(f"{u['prenom']} {u['nom']}") in t or t in _normalize(f"{u['prenom']} {u['nom']}"):
            return u['id']
    return None

@app.route('/taches/import', methods=['GET', 'POST'])
@login_required
def taches_import():
    if not current_user.is_staff: abort(403)
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename:
            flash('Aucun fichier sélectionné.', 'danger')
            return redirect(request.url)
        try:
            rows = parse_taches_file(f)
            if not rows:
                flash('Aucune tâche exploitable trouvée dans le fichier.', 'warning')
                return redirect(request.url)

            conn = get_db()
            existing = {(r['titre'], r['categorie']) for r in conn.execute('SELECT titre,categorie FROM taches').fetchall()}
            for r in rows:
                r['doublon'] = (r['titre'], r['categorie']) in existing
                r['responsable_id'] = _match_responsable(conn, r['responsable_txt'])
            conn.close()

            tmp_path = os.path.join(TMP_DIR, f'taches_{uuid.uuid4().hex}.json')
            with open(tmp_path, 'w', encoding='utf-8') as fp:
                json.dump(rows, fp, ensure_ascii=False)

            nb_doublons = sum(1 for r in rows if r['doublon'])
            cats = sorted(set(r['categorie'] for r in rows))
            return render_template('taches_import_preview.html', rows=rows, total=len(rows),
                                   nb_doublons=nb_doublons, cats=cats, tmp_file=os.path.basename(tmp_path))
        except Exception as e:
            flash(f'Erreur de lecture : {e}', 'danger')
    return render_template('taches_import.html')

@app.route('/taches/import/confirm', methods=['POST'])
@login_required
def taches_import_confirm():
    if not current_user.is_staff: abort(403)
    tmp_name = request.form.get('tmp_file', '')
    tmp_path = os.path.join(TMP_DIR, tmp_name)
    if not tmp_name or not os.path.exists(tmp_path):
        flash('Session expirée. Veuillez relancer l\'import.', 'danger')
        return redirect(url_for('taches_import'))
    with open(tmp_path, encoding='utf-8') as fp:
        rows = json.load(fp)
    os.unlink(tmp_path)

    skip_doublons = request.form.get('skip_doublons') == '1'
    conn = get_db()
    ok, skip = 0, 0
    for r in rows:
        if not can_edit(r['categorie']):
            skip += 1
            continue
        if r.get('doublon') and skip_doublons:
            skip += 1
            continue
        conn.execute('''INSERT INTO taches
            (titre,description,statut,priorite,echeance,categorie,responsable_id,delai_libelle,lien,created_by)
            VALUES (?,?,?,?,?,?,?,?,?,?)''',
            (r['titre'], r.get('notes',''), r['statut'], r['priorite'], r['echeance'],
             r['categorie'], r.get('responsable_id'), r['delai_libelle'], r.get('lien',''), current_user.id))
        ok += 1
    conn.commit(); conn.close()
    flash(f'✅ Import terminé — {ok} tâche(s) créée(s), {skip} ignorée(s).', 'success' if ok else 'warning')
    return redirect(url_for('taches_list'))

# ── BUDGET ────────────────────────────────────────────────────────────────────
def _budget_categories(conn):
    return [r['categorie'] for r in conn.execute(
        'SELECT DISTINCT categorie FROM budget_lignes ORDER BY categorie').fetchall()]

def _editable_budget_categories(conn):
    if current_user.is_admin:
        return _budget_categories(conn)
    return current_user.categories()

@app.route('/budget')
@login_required
def budget_list():
    if not current_user.is_staff: abort(403)
    conn = get_db()
    lignes = conn.execute('SELECT * FROM budget_lignes ORDER BY categorie, code').fetchall()
    by_cat = {}
    for l in lignes:
        by_cat.setdefault(l['categorie'], []).append(l)
    subtotals = {}
    for cat, items in by_cat.items():
        subtotals[cat] = {
            'cout_estime': sum(l['cout_estime'] for l in items),
            'subvention_prevue': sum(l['subvention_prevue'] for l in items),
            'reste_a_charge_prevu': sum(l['reste_a_charge_prevu'] for l in items),
            'cout_reel': sum(l['cout_reel'] for l in items),
            'subvention_recue': sum(l['subvention_recue'] for l in items),
        }
    totals = {
        'cout_estime': sum(l['cout_estime'] for l in lignes),
        'subvention_prevue': sum(l['subvention_prevue'] for l in lignes),
        'reste_a_charge_prevu': sum(l['reste_a_charge_prevu'] for l in lignes),
        'cout_reel': sum(l['cout_reel'] for l in lignes),
        'subvention_recue': sum(l['subvention_recue'] for l in lignes),
    }
    conn.close()
    return render_template('budget_list.html', by_cat=by_cat, subtotals=subtotals, totals=totals)

@app.route('/budget/new', methods=['GET','POST'])
@login_required
def budget_new():
    if not current_user.is_staff: abort(403)
    conn = get_db()
    editable_cats = _editable_budget_categories(conn)
    if request.method == 'POST':
        f = request.form
        categorie = f.get('categorie','').strip()
        if not can_edit(categorie):
            conn.close(); abort(403)
        cout = float(f.get('cout_estime') or 0)
        subv = float(f.get('subvention_prevue') or 0)
        conn.execute('''INSERT INTO budget_lignes
            (categorie,code,poste,detail,cout_estime,subvention_prevue,reste_a_charge_prevu,
             cout_reel,subvention_recue,statut,notes,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,datetime('now'))''',
            (categorie, f.get('code','').strip(), f['poste'].strip(), f.get('detail','').strip(),
             cout, subv, cout - subv,
             float(f.get('cout_reel') or 0), float(f.get('subvention_recue') or 0),
             f.get('statut','prevu'), f.get('notes','').strip()))
        conn.commit(); conn.close()
        flash('Ligne de budget créée.', 'success')
        return redirect(url_for('budget_list'))
    conn.close()
    return render_template('budget_form.html', action='new', item=None, editable_cats=editable_cats)

@app.route('/budget/<int:id>/edit', methods=['GET','POST'])
@login_required
def budget_edit(id):
    if not current_user.is_staff: abort(403)
    conn = get_db()
    item = conn.execute('SELECT * FROM budget_lignes WHERE id=?', (id,)).fetchone()
    if not item: conn.close(); abort(404)
    if not can_edit(item['categorie']):
        conn.close(); abort(403)
    editable_cats = _editable_budget_categories(conn)
    if request.method == 'POST':
        f = request.form
        categorie = f.get('categorie','').strip()
        if not can_edit(categorie):
            conn.close(); abort(403)
        cout = float(f.get('cout_estime') or 0)
        subv = float(f.get('subvention_prevue') or 0)
        conn.execute('''UPDATE budget_lignes SET categorie=?,code=?,poste=?,detail=?,cout_estime=?,
            subvention_prevue=?,reste_a_charge_prevu=?,cout_reel=?,subvention_recue=?,statut=?,notes=?,
            updated_at=datetime('now') WHERE id=?''',
            (categorie, f.get('code','').strip(), f['poste'].strip(), f.get('detail','').strip(),
             cout, subv, cout - subv,
             float(f.get('cout_reel') or 0), float(f.get('subvention_recue') or 0),
             f.get('statut','prevu'), f.get('notes','').strip(), id))
        conn.commit(); conn.close()
        flash('Ligne de budget mise à jour.', 'success')
        return redirect(url_for('budget_list'))
    conn.close()
    return render_template('budget_form.html', action='edit', item=item, editable_cats=editable_cats)

@app.route('/budget/<int:id>/delete', methods=['POST'])
@login_required
def budget_delete(id):
    if not current_user.is_staff: abort(403)
    conn = get_db()
    item = conn.execute('SELECT categorie FROM budget_lignes WHERE id=?', (id,)).fetchone()
    if not item: conn.close(); abort(404)
    if not can_edit(item['categorie']):
        conn.close(); abort(403)
    conn.execute('DELETE FROM budget_lignes WHERE id=?', (id,))
    conn.commit(); conn.close()
    flash('Ligne de budget supprimée.', 'success')
    return redirect(url_for('budget_list'))

# ── IMPORT BUDGET ─────────────────────────────────────────────────────────────
_CAT_RE = re.compile(r'^\s*\d+\s*-\s*')
_SUB_RE = re.compile(r'^\d+\.\d+$')

def parse_budget_file(file_storage):
    import openpyxl
    wb = openpyxl.load_workbook(file_storage, data_only=True)
    ws = None
    for s in wb.worksheets:
        if 'budget' in _normalize(s.title):
            ws = s
            break
    if ws is None:
        ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    data = []
    current_cat = ''
    for row in rows[8:]:
        if not row or row[0] is None:
            continue
        code_s = str(row[0]).strip()
        if code_s.upper().startswith('TOTAL'):
            continue
        if _CAT_RE.match(code_s):
            current_cat = _CAT_RE.sub('', code_s).strip()
            continue
        if not _SUB_RE.match(code_s):
            continue
        poste = row[1] if len(row) > 1 else None
        if not poste:
            continue  # placeholder empty sub-line
        detail = row[2] if len(row) > 2 and row[2] else ''
        cout = float(row[3]) if len(row) > 3 and row[3] else 0.0
        subv = float(row[4]) if len(row) > 4 and row[4] else 0.0
        reste = float(row[5]) if len(row) > 5 and row[5] is not None else (cout - subv)
        data.append(dict(categorie=current_cat, code=code_s, poste=str(poste).strip(),
                          detail=str(detail).strip(), cout_estime=cout, subvention_prevue=subv,
                          reste_a_charge_prevu=reste))
    return data

@app.route('/budget/import', methods=['GET', 'POST'])
@login_required
def budget_import():
    if not current_user.is_admin: abort(403)
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename:
            flash('Aucun fichier sélectionné.', 'danger')
            return redirect(request.url)
        try:
            rows = parse_budget_file(f)
            if not rows:
                flash('Aucune ligne de budget exploitable trouvée dans le fichier.', 'warning')
                return redirect(request.url)

            conn = get_db()
            existing = {(r['categorie'], r['code']) for r in conn.execute('SELECT categorie,code FROM budget_lignes').fetchall()}
            conn.close()
            for r in rows:
                r['doublon'] = (r['categorie'], r['code']) in existing

            tmp_path = os.path.join(TMP_DIR, f'budget_{uuid.uuid4().hex}.json')
            with open(tmp_path, 'w', encoding='utf-8') as fp:
                json.dump(rows, fp, ensure_ascii=False)

            nb_doublons = sum(1 for r in rows if r['doublon'])
            total_estime = sum(r['cout_estime'] for r in rows)
            total_subv = sum(r['subvention_prevue'] for r in rows)
            return render_template('budget_import_preview.html', rows=rows, total=len(rows),
                                   nb_doublons=nb_doublons, total_estime=total_estime, total_subv=total_subv,
                                   tmp_file=os.path.basename(tmp_path))
        except Exception as e:
            flash(f'Erreur de lecture : {e}', 'danger')
    return render_template('budget_import.html')

@app.route('/budget/import/confirm', methods=['POST'])
@login_required
def budget_import_confirm():
    if not current_user.is_admin: abort(403)
    tmp_name = request.form.get('tmp_file', '')
    tmp_path = os.path.join(TMP_DIR, tmp_name)
    if not tmp_name or not os.path.exists(tmp_path):
        flash('Session expirée. Veuillez relancer l\'import.', 'danger')
        return redirect(url_for('budget_import'))
    with open(tmp_path, encoding='utf-8') as fp:
        rows = json.load(fp)
    os.unlink(tmp_path)

    skip_doublons = request.form.get('skip_doublons') == '1'
    conn = get_db()
    ok, skip = 0, 0
    for r in rows:
        if r.get('doublon') and skip_doublons:
            skip += 1
            continue
        conn.execute('''INSERT INTO budget_lignes
            (categorie,code,poste,detail,cout_estime,subvention_prevue,reste_a_charge_prevu,updated_at)
            VALUES (?,?,?,?,?,?,?,datetime('now'))''',
            (r['categorie'], r['code'], r['poste'], r['detail'],
             r['cout_estime'], r['subvention_prevue'], r['reste_a_charge_prevu']))
        ok += 1
    conn.commit(); conn.close()
    flash(f'✅ Import terminé — {ok} ligne(s) créée(s), {skip} ignorée(s).', 'success' if ok else 'warning')
    return redirect(url_for('budget_list'))

# ── EXPORT TÂCHES ────────────────────────────────────────────────────────────
STATUT_LABELS = {'todo': 'À faire', 'en_cours': 'En cours', 'termine': 'Terminé'}

from excel_export import build_export_workbook  # à ajouter en haut du fichier avec tes autres imports

@app.route('/taches/export')
@login_required
def taches_export():
    if not current_user.is_staff:
        abort(403)
    conn = get_db()
    taches = conn.execute('''
        SELECT t.*, u.nom resp_nom, u.prenom resp_prenom
        FROM taches t
        LEFT JOIN users u ON t.responsable_id = u.id
        ORDER BY t.categorie, t.created_at
    ''').fetchall()

    headers = ['N°', 'Catégorie', 'Titre', 'Délai', 'Échéance', 'Priorité',
               'Responsable', 'Statut', 'Lien', 'Notes']

    rows = []
    for i, t in enumerate(taches, start=1):
        responsable = f"{t['resp_prenom']} {t['resp_nom']}" if t['resp_nom'] else ''
        rows.append({
            'n': str(i),
            'categorie': t['categorie'] or '',
            'titre': t['titre'],
            'delai': t['delai_libelle'] or '',
            'echeance': t['echeance'] or '',
            'priorite': t['priorite'] or '',
            'responsable': responsable,
            'statut': STATUT_LABELS.get(t['statut'], t['statut']),
            'lien': t['lien'] or '',
            'notes': t['description'] or '',
        })

        sous_taches = conn.execute(
            'SELECT * FROM sous_taches WHERE tache_id=? ORDER BY created_at', (t['id'],)
        ).fetchall()
        for j, st in enumerate(sous_taches, start=1):
            rows.append({
                'n': f'{i}.{j}',
                'categorie': t['categorie'] or '',
                'titre': f"      ↳ {st['titre']}",
                'delai': '',
                'echeance': '',
                'priorite': '',
                'responsable': '',
                'statut': 'Fait' if st['statut'] == 'fait' else 'À faire',
                'lien': st['lien'] or '',
                'notes': '',
            })

    conn.close()

    wb = build_export_workbook(
        export_name="Tâches",
        headers=headers,
        rows=rows,
        category_field="categorie",
        priority_field="priorite",
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                      download_name=f'taches_dole2028_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── EXPORT PLANNING ───────────────────────────────────────────────────────────
@app.route('/planning/export')
@login_required
def planning_export():
    if not current_user.is_staff:
        abort(403)
    conn = get_db()
    missions = conn.execute('''
        SELECT m.*, pm.nom parent_nom
        FROM missions m
        LEFT JOIN missions pm ON m.parent_id = pm.id
        ORDER BY m.jour, m.heure_debut
    ''').fetchall()

    headers = ['Jour', 'Mission', 'Site', 'Début', 'Fin', 'Places',
               'Affectés', 'Manque', 'Bénévoles', 'Sous-mission de']

    rows = []
    for m in missions:
        aff = conn.execute('''
            SELECT u.nom, u.prenom FROM affectations a
            JOIN benevoles b ON a.benevole_id = b.id
            JOIN users u ON b.user_id = u.id
            WHERE a.mission_id=? ORDER BY u.nom
        ''', (m['id'],)).fetchall()
        noms = ', '.join(f"{a['prenom']} {a['nom']}" for a in aff)
        rows.append({
            'jour': m['jour'],
            'mission': m['nom'],
            'site': m['site'],
            'debut': m['heure_debut'],
            'fin': m['heure_fin'],
            'places': m['nb_places'],
            'affectes': len(aff),
            'manque': max(m['nb_places'] - len(aff), 0),
            'benevoles': noms,
            'parent': m['parent_nom'] or '',
        })
    conn.close()

    wb = build_export_workbook(
        export_name="Planning bénévoles",
        headers=headers,
        rows=rows,
        category_field="jour",
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                      download_name=f'planning_dole2028_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── RUN ───────────────────────────────────────────────────────────────────────
init_db()
migrate_db()

if __name__ == '__main__':
    init_db()
    print("\n" + "="*52)
    print("   DOLE 2028 — Application de Gestion FSCF")
    print("="*52)
    print("   Navigateur  : http://localhost:5000")
    print("   Admin       : admin@dole2028.fr")
    print("   Mot de passe: admin2028")
    print("="*52 + "\n")
    app.run(debug=True, host='0.0.0.0', port=5000)
 
